import datetime
import csv
import sys
import os
import numpy as np
import torch
import torch.nn as nn
import math
import matplotlib as mpl
import matplotlib.pyplot as plt
from models.resnet import resnet18, resnet34, resnet50, resnet101, resnet152
from models.vgg import vgg16
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

def get_load_weight_path(args, load_weight_path=None):
    if load_weight_path is None:
        load_weight_path = 'weights/pretrained/' + args.model + '.pth'
    print("load weight path: ", load_weight_path)
    return load_weight_path


def get_save_weight_path(args):
    date = datetime.datetime.now().strftime("%y%m%d")
    time = datetime.datetime.now().strftime("%H%M")
    num_bits = '' if args.quant == 'float' else str(args.fixed_num)
    fn = '' if args.quant == 'float' else str(args.fn)
    if args.quant == 'fixed':
        quant = 'fix'
    elif args.quant == 'bit_serial':
        quant = 'ser'
    elif args.quant == 'float':
        quant = 'flo'
    else:
        quant = ''

    high = 'h%s' % str(args.high_bits) if args.high_bits else ''
    # model = args.model + '_' + args.dataset + quant + num_bits + fn
    if args.quant == 'float':
        model = '%s_%s_%s' % (args.model, args.dataset, quant)
    elif args.quant == 'bit_serial':
        if args.diff_fn:
            if args.high_bits:
                model = '%s_%s_%s_%s_%s' % (args.model, args.dataset, quant, num_bits, high)
            else:
                model = '%s_%s_%s_%s' % (args.model, args.dataset, quant, num_bits)
        else:
            if args.high_bits:
                model = '%s_%s_%s_%s_%s_%s' % (args.model, args.dataset, quant, num_bits, fn, high)
            else:
                model = '%s_%s_%s_%s_%s' % (args.model, args.dataset, quant, num_bits, fn)
    # save_weight_path = 'weights/' + model + '_' + date + '/'    
    save_weight_path = 'weights/%s/%s_%s/' % (model, model, date)
    # fb = '_fb' if args.freeze_bn else ''
    # file_name = model + high + '_' + date + '_' + time + '.pth'
    file_name = '%s_%s_%s.pth' % (model, date, time)
    print("save file name  : ", file_name)
    return save_weight_path, file_name, model


def get_model(args, device, settings):
    # get num_classes
    if args.dataset == 'cifar10':
        num_classes = 10
    elif args.dataset == 'cifar100':
        num_classes = 100
    elif args.dataset == 'imagenet':
        num_classes = 1000
    else:
        print("#### unknown dataset ####")
        sys.exit()

    model = args.model

    return eval(model)(settings=settings,
                       quant=args.quant,
                       num_classes=num_classes,
                       fixed_num=args.fixed_num,
                       fract_num=args.fn,
                       high_bits=args.high_bits,
                       diff_fn=args.diff_fn).to(device)


def csv_write(filename, model_name, acc_max, train_batch, test_batch, num_worker, num_epochs, 
                lr, step_size, gamma, fixed_num, data_select):
    with open(filename, 'a') as f:
        writer = csv.writer(f)
        # writer.writerow([*args])
        writer.writerow(['model_name: {}, data_select: {}, acc_max: {}'.format(model_name, data_select, acc_max)])
        writer.writerow(['train_batch: {}, test_batch: {}, num_worker: {}, epoch: {}'.format(train_batch, test_batch, num_worker, num_epochs)])
        writer.writerow(['LR: {}, step_size: {}, gamma: {}, fixed_num: {}'.format(lr, step_size, gamma, fixed_num)])


def add_param(writer, net, step):
    for name, value in net.named_parameters():
        writer.add_histogram(name, value, step)
    
def dist_loadedweght(args, model):
    layername = ["Conv1_1", "Conv1_2", "Conv2_1", "Conv2_2", "Conv3_1", "Conv3_2", "Conv3_3", "Conv4_1", 
    "Conv4_2", "Conv4_3", "Conv5_1", "Conv5_2", "Conv5_3", "Linear1", "Linear2", "Linear3"]
    l = 0
    min_row = 2 
    load_weight_path = args.weights
    name = os.path.splitext(os.path.basename(load_weight_path))[0]
    path = '/artic/t-kaneko/Documents/progressiveNN/%s' % name
    os.mkdir(path)

    wb = Workbook()
    # ws = wb.worksheets[0]
    for m in model.modules():
        if isinstance(m, nn.Conv2d) or isinstance(m, nn.Linear):
            # filename = "%s.xlsx" % layername[l]
            filename = "weight_dist.xlsx"
            # ws = wb.worksheets[l]
            ws = wb.create_sheet()
            xlfile = os.path.join(path, filename)
            weight = m.weight.cpu().detach().numpy().flatten()
            value ,count = np.unique(weight,return_counts=True)
            len_data = len(value)
            max_row = min_row + len_data - 1

            # ws.cell(row = 1, column = 1 , value = layername[l])
            ws.cell(row = 1, column = 1 , value = "value")
            ws.cell(row = 1, column = 2, value = "count")

            for i in range(0, len_data):
                ws.cell(row = i + 2, column = 1, value = value[i])
                ws.cell(row = i + 2, column = 2, value = count[i])
            x_data = Reference(ws, min_col = 1, min_row = min_row, max_row = max_row)
            y_data = Reference(ws, min_col = 2, min_row = min_row, max_row = max_row)
            series = Series(y_data, x_data)
            chart = ScatterChart()
            chart.title = layername[l]
            chart.legend = None
            chart.series.append(series)
            ws.add_chart(chart)
            wb.save(xlfile)
            l += 1


def mahalanobis_distance(x):
    distance = ((x.max() - x.mean()).abs()) / x.std()
    # print(distance)
    if 0 < distance <= 2.2745:
        ideal_bits = 8
    elif 2.2745 < distance <= 3.1212:
        ideal_bits = 7
    elif 3.1212 < distance <= 3.5365:
        ideal_bits = 6
    elif 3.5365 < distance <= 3.8788:
        ideal_bits = 5
    elif 3.8788 < distance <= 4.2114:
        ideal_bits = 4
    elif 4.2114 < distance <= 4.6143:
        ideal_bits = 3
    elif 4.6143 < distance <= 3.1448:
        ideal_bits = 2
    elif 3.1448 < distance <= 3.9574:
        ideal_bits = 1
    # print(ideal_bits)
    return ideal_bits


def entropy_threshold(x):
    entropy = 0
    for xx in x:
        entropy += - xx * torch.log2(xx)
    if 0 < entropy <= 0.89262 :
        ideal_bits = 1
    # elif 0.89262 < entropy <= 5.59707:
    #     ideal_bits = 2
    elif 0.89262 < entropy <= 1.06668:
        ideal_bits = 3
    elif 1.06668 < entropy:
        ideal_bits = 4
    # elif 3.8788 < entropy <= 4.2114:
    #     ideal_bits = 5
    # elif 4.2114 < entropy <= 4.6143:
    #     ideal_bits = 6
    # elif 4.6143 < entropy <= 6.3183:
    #     ideal_bits = 7
    # elif 3.1448 < entropy <= 3.9574:
    #     ideal_bits = 8
    # print(ideal_bits)
    return ideal_bits


def csv_save(x, title):
    basename = os.path.splitext(os.path.basename(title))[0]
    date = datetime.datetime.now().strftime("%y%m%d")
    time = datetime.datetime.now().strftime("%H%M")
    filepath = "csv/%s" % date
    os.makedirs(filepath, exist_ok=True)
    filename = 'csv/%s/%s_%s_%s.csv' % (date, date, time, basename)
    x.to_csv(filename)


def plt_cum(y1, y2, y3, high_bits, ylabel, data_select, name, key, ratio, e_th, t_th):
    fig = plt.figure()
    # ax1 = fig.add_subplot(111)
    # ax1.plot(y1, label="correct")
    # ax1.plot(y2, label="incorrect")
    # ax1.set_label('%s' % ylabel)
    # ax1.set_yticks(np.linspace(0, 50000, 5))
    # ax1.set_ylim(-2500, 52500)
    # ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, loc: "{:,}".format(int(x))))
    # ax1.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, loc: "{:,}".format(int(x))))
    
    # ax2 = ax1.twinx()
    ax2 = fig.add_subplot(111, xlabel='Image number', ylabel='correct_rate')
    ax2.plot(key, ratio, color='black', marker='.')
    ax2.text(25000, 0.9998, "({}, {:.5f}, entropy: {:.5f})".format(key, ratio, e_th))
    ax2.plot(y3, color='m', label="correct rate")
    # ax2.set_yticks(np.linspace(0, 1, 5))
    # ax2.set_ylim(-0.05, 1.05)
    
    # ax2.hlines([0.99], 0, 50000, 'green', 'dashed')
    ax2.set_label('correct ratio')
    fig.legend(loc='upper left')
    ax2.set_title('resnet18,%sbit,cifar100,correct_rate_only' % high_bits)
    basename = os.path.splitext(os.path.basename(name))[0]
    date = datetime.datetime.now().strftime("%y%m%d")
    time = datetime.datetime.now().strftime("%H%M")
    filepath = "plt/%s/%s/%s" % (ylabel, data_select, date)
    os.makedirs(filepath, exist_ok=True)
    plt.savefig("plt/%s/%s/%s/%s_%s_%s_val%s_%s.png" % (ylabel, data_select, date, date, time, basename, int(100*t_th), ylabel))
    plt.close()

def plt_number_entropy(x, high_bits, ylabel, data_select, name):
    fig = plt.figure()
    ax = fig.add_subplot(111, xlabel='Image number', ylabel=ylabel)
    ax.set_title('resnet18,%sbit,cifar100,correct_rate_only' % high_bits)
    correct = x[0] * x[1]
    incorrect = x[0] * x[2]
    # breakpoint()
    ax.plot((np.where(correct != 0)[0]), correct[correct != 0], label='correct')
    ax.plot((np.where(incorrect != 0)[0]),incorrect[incorrect != 0], marker='.', markersize=0.1, label='incorrect')

    fig.legend()
    basename = os.path.splitext(os.path.basename(name))[0]
    date = datetime.datetime.now().strftime("%y%m%d")
    time = datetime.datetime.now().strftime("%H%M")
    filepath = "plt/%s/%s/%s" % (ylabel, data_select, date)
    os.makedirs(filepath, exist_ok=True)
    plt.savefig("plt/%s/%s/%s/%s_%s_%s_val_%s.png" % (ylabel, data_select, date, date, time, basename, ylabel))
    plt.close()

def correct_rate_entropy(correct_rate, entropy, high_bits, name, ylabel, data_select):
    fig = plt.figure()
    ax1 = fig.add_subplot(111, xlabel='Image number', ylabel='correct rate')
    ax1.plot(correct_rate, label="correct rate")
    ax1.set_yticks(np.linspace(0, 1, 5))
    ax1.set_ylim(-0.05, 1.05)

    ax2 = ax1.twinx()
    ax2.plot(entropy, color='orange', label='entropy')
    ax2.set_ylabel('entropy')
    ax1.set_title('resnet18, %sbit, cifar100' % high_bits)
    fig.legend()
    basename = os.path.splitext(os.path.basename(name))[0]
    date = datetime.datetime.now().strftime("%y%m%d")
    time = datetime.datetime.now().strftime("%H%M")
    filepath = "plt/%s/%s/%s" % (ylabel, data_select, date)
    os.makedirs(filepath, exist_ok=True)
    plt.savefig("plt/%s/%s/%s/%s_%s_%s_val.png" % (ylabel, data_select, date, date, time, basename))
    plt.close()

def numpy_save(x, data_select, name, flag):
    basename = os.path.splitext(os.path.basename(name))[0]
    date = datetime.datetime.now().strftime("%y%m%d")
    time = datetime.datetime.now().strftime("%H%M")
    filepath = "numpy/%s/%s/%s" % (flag, data_select, date)
    os.makedirs(filepath, exist_ok=True)
    filename = '%s_%s_%s' % (date, time, basename)
    np.save('%s/%s' % (filepath, filename), x)

    return filename


def loss_bit(entropy, en_sort_list):
    loss = np.zeros([len(en_sort_list)])
    for (i, l) in enumerate(en_sort_list):
        loss[i] = 1 - l[1][np.where(l[0] > entropy)[0][0]]
    return loss

def threshold(alpha, en_sort_list):
    a = 1
    b = 0
    x = (np.arange(len(en_sort_list)) + 1)
    comp = a * x + b
    resol =  300  # resolution
    en_max = 6  # max entropy
    en_list = np.arange(1, resol+1) / resol * en_max
    scale_factor = alpha
    th_en = np.zeros([len(x)])

    for e in en_list:
        loss = loss_bit(e, en_sort_list)
        obj_func = loss + comp * scale_factor
        th_en[obj_func.argmin()] = e
    return th_en[:3]

